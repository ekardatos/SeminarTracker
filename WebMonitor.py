import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime

# Base URL of the website being scraped
BASE_URL = "https://ekek.gr"

# Full URL of the seminars listing page
LISTING_URL = "https://ekek.gr/seminaria/"

# Name of the Excel file where results are saved
EXCEL_FILE = "seminars.xlsx"


def get_seminar_links(listing_url):
    """
    Fetches the seminar listing page and returns a list of URLs
    for seminars organized by the University of Patras.
    """
    # Request the listing page
    response = requests.get(listing_url)

    # Stop early if the page couldn't be reached
    if response.status_code != 200:
        print("Failed to retrieve the listing.")
        return []

    # Parse the HTML content of the page
    soup = BeautifulSoup(response.text, 'html.parser')

    # List to collect valid seminar URLs
    seminar_links = []

    # Find all anchor tags that point to a seminar page
    seminar_items = soup.select("a[href*='/seminaria/']")

    for seminar_item in seminar_items:
        # Look for the organizer label div inside the seminar item
        organizer_label = seminar_item.find('div', string=lambda t: t and 'Διοργανωτ' in t)

        if organizer_label:
            # Get the sibling div that contains the actual organizer name
            organizer_value = organizer_label.find_next_sibling('div')

            # Only keep seminars from the University of Patras
            if organizer_value and 'Πανεπιστήμιο Πατρών' in organizer_value.get_text(strip=True):
                href = seminar_item.get('href')

                # Skip empty or anchor-only links
                if href and not href.startswith('#'):

                    # Convert relative URLs to absolute
                    if not href.startswith('http'):
                        href = f"{BASE_URL}{href}"

                    seminar_links.append(href)

    return seminar_links


def scrape_seminar_page(seminar_url, seminar_titles, seminar_details):
    """
    Visits a single seminar page and extracts the title and sidebar details.
    Appends the results directly to the provided lists.
    """
    # Request the individual seminar page
    response = requests.get(seminar_url)

    if response.status_code == 200:
        # Parse the page HTML
        soup = BeautifulSoup(response.text, 'html.parser')

        # Extract the seminar title from the first <h1> tag
        seminar_name_tag = soup.find('h1')
        seminar_name = seminar_name_tag.get_text(strip=True) if seminar_name_tag else "No seminar name found"
        seminar_titles.append(seminar_name)

        # Look for the sidebar div that contains the seminar details
        sidebar_div = soup.find('div', class_='sidebar__inner gap')

        if sidebar_div:
            print(f"[✓] Found details for: {seminar_name}")

            # Extract all text from the sidebar as a single string
            raw_text = sidebar_div.get_text(separator=" ", strip=True)
            seminar_details.append(raw_text)
        else:
            print(f"[!] Sidebar not found in: {seminar_url}")
            seminar_details.append("No details found.")
    else:
        print(f"[✗] Failed to retrieve seminar page at {seminar_url} (status code: {response.status_code})")


def save_to_excel(seminar_titles, seminar_details):
    """
    Saves the scraped seminar data to an Excel file.
    If a seminar already exists, it checks for changes and logs updates.
    If it's new, it adds a fresh row with the current date.
    """
    # Try to open an existing workbook, or create a new one if not found
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Seminars"

        # Write column headers on first run
        sheet["B1"] = "Seminar Title"
        sheet["C1"] = "Date of Execution"
        sheet["D1"] = "Seminar Details"
        sheet["F1"] = "Date of Update"
        sheet["G1"] = "New Seminar Details"
        sheet["H1"] = "Date of Update 2"
        sheet["I1"] = "New Seminar Details 2"

    # Capture the current timestamp for this run
    execution_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Read all existing seminar titles from column B for comparison
    existing_titles = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]

    for i in range(len(seminar_titles)):
        title = seminar_titles[i]
        details = seminar_details[i]

        if title in existing_titles:
            # Find the row index of the existing seminar
            index = existing_titles.index(title)
            row = index + 2  # Offset by 2: 1 for header row, 1 for 0-based index

            # Start from column D (initial details) and find the next empty update slot
            last_column = 4
            while sheet.cell(row=row, column=last_column).value is not None:
                last_column += 2  # Each update occupies 2 columns: date + details

            # Get the most recently stored details for comparison
            last_details = sheet.cell(row=row, column=last_column - 1).value

            # Only log an update if the details have actually changed
            if details != last_details:
                sheet.cell(row=row, column=last_column, value=execution_date)
                sheet.cell(row=row, column=last_column + 1, value=details)
        else:
            # Seminar is new — add it as a fresh row at the bottom
            new_row = sheet.max_row + 1
            sheet.cell(row=new_row, column=2, value=title)
            sheet.cell(row=new_row, column=3, value=execution_date)
            sheet.cell(row=new_row, column=4, value=details)

    # Persist all changes to disk
    wb.save(EXCEL_FILE)
    print(f"Data processing complete. Results saved to '{EXCEL_FILE}'.")


def main():
    """
    Entry point of the script.
    Coordinates the full flow: fetch links → scrape pages → save to Excel.
    """
    # Step 1: Get all seminar links from the listing page
    seminar_links = get_seminar_links(LISTING_URL)

    if seminar_links:
        print(f"\nFound {len(seminar_links)} seminars from University of Patras. Scraping details...\n")

        # Step 2: Initialize empty lists to collect results
        seminar_titles = []
        seminar_details = []

        # Step 3: Scrape each seminar page and populate the lists
        for link in seminar_links:
            scrape_seminar_page(link, seminar_titles, seminar_details)

        # Step 4: Save all collected data to Excel
        save_to_excel(seminar_titles, seminar_details)
    else:
        print("No seminars found for University of Patras.")


if __name__ == "__main__":
    main()